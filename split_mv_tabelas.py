"""
split_mv_tabelas.py
Separa o MV Tabelas.xlsx em multiplos arquivos Excel categorizados por area funcional.

Uso:
    python split_mv_tabelas.py [--input "MV Tabelas.xlsx"] [--output "output_mv_categorizado"]

Dependencias:
    pip install pandas openpyxl
"""

import re
import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CATEGORIAS
# ============================================================
CATEGORIES = {
    '01_CADASTRO_PACIENTE': {
        'name': 'Cadastro e Paciente',
        'description': 'Cadastro de pacientes, dados demograficos, convenios e vinculacoes',
    },
    '02_ATENDIMENTO': {
        'name': 'Atendimento e Internacao',
        'description': 'Atendimentos, internacoes, altas, leitos, transferencias',
    },
    '03_AMBULATORIO_AGENDAMENTO': {
        'name': 'Ambulatorio e Agendamento',
        'description': 'Consultas ambulatoriais, agenda, central de marcacao',
    },
    '04_URGENCIA_CLASSIFICACAO_RISCO': {
        'name': 'Urgencia e Classificacao de Risco',
        'description': 'Pronto-socorro, triagem, classificacao Manchester (SACR)',
    },
    '05_PRONTUARIO_ELETRONICO': {
        'name': 'Prontuario Eletronico (PEP)',
        'description': 'Documentos clinicos, evolucoes, PW_ (PEP Web)',
    },
    '06_PRESCRICAO_MEDICACAO': {
        'name': 'Prescricao e Medicacao',
        'description': 'Prescricoes, itens, aprazamento, checagem, CPOE (MVCPOE)',
    },
    '07_CENTRO_CIRURGICO': {
        'name': 'Centro Cirurgico',
        'description': 'Cirurgias, agendamento cirurgico, anestesia, OPME',
    },
    '08_EXAMES_LABORATORIO': {
        'name': 'Exames e Laboratorio',
        'description': 'Pedidos, resultados, laudos, SADT',
    },
    '09_IMAGEM_PACS': {
        'name': 'Imagens Medicas e PACS',
        'description': 'DICOM, PACS, radiologia (ORDDCM_, MPACS)',
    },
    '10_FARMACIA_ESTOQUE': {
        'name': 'Farmacia e Estoque',
        'description': 'Produtos, movimentacao, almoxarifado, solicitacoes',
    },
    '11_COMPRAS_CONTRATOS': {
        'name': 'Compras e Contratos',
        'description': 'Pedidos de compra, fornecedores, contratos (GCTO_), licitacoes',
    },
    '12_FATURAMENTO_CONTAS': {
        'name': 'Faturamento e Contas Medicas',
        'description': 'Contas, procedimentos, valores, glosas',
    },
    '13_FINANCEIRO': {
        'name': 'Financeiro',
        'description': 'Contas a pagar/receber, caixa, tesouraria, titulos',
    },
    '14_REGULATORIO_TISS_SUS': {
        'name': 'Regulatorio TISS e SUS',
        'description': 'TISS, ANS, AIH, APAC, SPED, REINF, Nota Fiscal',
    },
    '15_PLANOS_SAUDE': {
        'name': 'Planos de Saude e Operadoras',
        'description': 'Gestao de planos, operadoras, autorizacoes (DBAPS)',
    },
    '16_PATRIMONIO_ENGENHARIA': {
        'name': 'Patrimonio e Engenharia Clinica',
        'description': 'Bens, equipamentos, manutencao, calibracao',
    },
    '17_BI_DATA_WAREHOUSE': {
        'name': 'BI e Data Warehouse',
        'description': 'Tabelas fato/dimensao, OLAP (MVCCADMIN, DBADW)',
    },
    '18_INTEGRACAO': {
        'name': 'Integracao e Interoperabilidade',
        'description': 'APIs, HL7, FUSION_, INTEGRA_, IMV_',
    },
    '19_PORTAL_DASHBOARD': {
        'name': 'Portal e Painel',
        'description': 'Portais, dashboards, portlets (DBACP)',
    },
    '20_SEGURANCA_AUDITORIA': {
        'name': 'Seguranca e Auditoria',
        'description': 'Usuarios, permissoes, logs, audit trail (DBASGU)',
    },
    '21_CONFIGURACAO_SISTEMA': {
        'name': 'Configuracao e Sistema',
        'description': 'Parametros, tipos, lookup tables, metadata',
    },
    '22_VIEWS_RELATORIOS': {
        'name': 'Views e Relatorios',
        'description': 'Views (VDIC_, V_, VW_), definicoes de relatorios',
    },
    '99_OUTROS': {
        'name': 'Outros',
        'description': 'Tabelas nao classificadas (temporarias, backup, misc)',
    },
}


# ============================================================
# REGRAS DE CLASSIFICACAO
# ============================================================

# Fase 1: Owner inteiro -> categoria
OWNER_RULES = {
    'DBAPS': '15_PLANOS_SAUDE',
    'DBASGU': '20_SEGURANCA_AUDITORIA',
    'MVCPOE': '06_PRESCRICAO_MEDICACAO',
    'MVCCADMIN': '17_BI_DATA_WAREHOUSE',
    'DBADW': '17_BI_DATA_WAREHOUSE',
    'DBACP': '19_PORTAL_DASHBOARD',
    'MVINTEGRA': '18_INTEGRACAO',
    'MPACS': '09_IMAGEM_PACS',
    'MPACSCRER': '09_IMAGEM_PACS',
    'IDCE': '09_IMAGEM_PACS',
    'IDCECRER': '09_IMAGEM_PACS',
    'AUDIT_DBAMV': '20_SEGURANCA_AUDITORIA',
    'AUDIT_DBASGU': '20_SEGURANCA_AUDITORIA',
    'AUDIT_DBAPS': '20_SEGURANCA_AUDITORIA',
    'HEPIC': '08_EXAMES_LABORATORIO',
    'FUSION': '18_INTEGRACAO',
    'FUSION_INTEGRACAO': '18_INTEGRACAO',
    'DBATUALIZA': '21_CONFIGURACAO_SISTEMA',
    'DBAMS': '21_CONFIGURACAO_SISTEMA',
    'GTP': '18_INTEGRACAO',
    'GTP_INTS': '18_INTEGRACAO',
    'HMED': '06_PRESCRICAO_MEDICACAO',
    'MVCONT': '13_FINANCEIRO',
    'EDITOR': '05_PRONTUARIO_ELETRONICO',
    'EPIMED': '08_EXAMES_LABORATORIO',
    'REMWEB': '18_INTEGRACAO',
    'APOIOINTEGRADORCLIENT': '18_INTEGRACAO',
    'FAROL': '19_PORTAL_DASHBOARD',
    'CONSULTA': '03_AMBULATORIO_AGENDAMENTO',
    'ACESSOMV': '20_SEGURANCA_AUDITORIA',
    'ACESSOPRD': '20_SEGURANCA_AUDITORIA',
    'GCMMV': '21_CONFIGURACAO_SISTEMA',
    'MVONEPASS': '20_SEGURANCA_AUDITORIA',
    'CARTORIO': '01_CADASTRO_PACIENTE',
    'PLANISA_DBREAD': '17_BI_DATA_WAREHOUSE',
    'FLOWTI': '21_CONFIGURACAO_SISTEMA',
    'ZABBIX_MON': '21_CONFIGURACAO_SISTEMA',
}

# Owners adicionais
OWNER_RULES.update({
    'DBAPORTAL': '19_PORTAL_DASHBOARD',
    'HUGOLMV': '21_CONFIGURACAO_SISTEMA',
    'VIVACE_SESGO_HUGOL': '19_PORTAL_DASHBOARD',
    'VIVACE_SESGO_CRER': '19_PORTAL_DASHBOARD',
    'SESGO': '19_PORTAL_DASHBOARD',
    'HGG': '21_CONFIGURACAO_SISTEMA',
    'MIGRATOR': '21_CONFIGURACAO_SISTEMA',
    'DBASW': '18_INTEGRACAO',
})

# Owners que vao para 99_OUTROS (Oracle internos, recovery, etc.)
OWNER_OTHERS = {
    'ORDDATA', 'PERFSTAT', 'GSMADMIN_INTERNAL', 'OJVMSYS', 'APPQOSSYS',
    'AUDSYS', 'SCOTT', 'DBSFWUSER', 'CLAUDIO_REIS',
    'MRRECOVERY', 'MRRECOVERY_HUGOL', 'MRRECOVERY_HECAD', 'MRRECOVERY_HDS',
}

# Fase 2: Prefixo do TABLE_NAME (regex compilado)
PREFIX_RULES = [
    # Views primeiro
    (re.compile(r'^(VDIC_|VW_|V_|VM_|VIEW)', re.IGNORECASE), '22_VIEWS_RELATORIOS'),

    # PACS/DICOM
    (re.compile(r'^(ORDDCM_|MPACS_|DCM_|DICOM_)', re.IGNORECASE), '09_IMAGEM_PACS'),

    # PEP Web / Prontuario
    (re.compile(r'^(PW_|PL_CUID|HT_PW|SAE_|SAE$|EDITOR_|EVOLUCAO|EVOL_)', re.IGNORECASE), '05_PRONTUARIO_ELETRONICO'),

    # Prescricao
    (re.compile(r'^(PRE_MED|PRE_PAD|ITPRE_|HRITPRE_|PRESCRICAO|ITEM_PRESCRICAO|HORARIO_ITEM_PRESC|HORARIO_CONSUMO|APRAZAMENTO)', re.IGNORECASE), '06_PRESCRICAO_MEDICACAO'),

    # Regulatorio TISS/SUS
    (re.compile(r'^(TISS_|PTU_|ANS_|APAC$|APAC_|AIH_|AIH$|BPA_|SPED_|REINF$|REINF_|NOTA_FISCAL|NFS_|DIRF$|DIRF_|DARF_|SEFIP_|RAIS_|CAGED_|ESOCIAL_|PROCEDIMENTO_SUS|RAAS|DCIH)', re.IGNORECASE), '14_REGULATORIO_TISS_SUS'),

    # Classificacao de risco
    (re.compile(r'^(SACR_|CLAS_RISCO|TRIAGEM|CLASS_RISCO)', re.IGNORECASE), '04_URGENCIA_CLASSIFICACAO_RISCO'),

    # Centro cirurgico
    (re.compile(r'^(CIRURGIA|AVISO_CIRURGIA|ITAVISO_|CIR_|AGE_CIR|AGENDA_CIR|EQUIPE_CIR|OPME_|OPMENEXO|ANEST)', re.IGNORECASE), '07_CENTRO_CIRURGICO'),

    # Exames/laboratorio
    (re.compile(r'^(RESULTADO_|PEDIDO_|MAPA_EXAME|LAB_|EXAME_|EXA_|ITPED_|LAUDO$|LAUDO_|SADT_|ATEND_SADT|AMOSTRA|SET_EXA|ITEM_COLETA|COLETA$|CULTURA$|CULTURAS$|ANTIMICROB|GERMES$|BANCADA$|LABORATOR)', re.IGNORECASE), '08_EXAMES_LABORATORIO'),

    # Farmacia/estoque
    (re.compile(r'^(ESTOQUE|MVTO_ESTOQUE|ITMVTO_ESTOQUE|ITMVTO_ENT|PRODUTO$|PRODUTO_|UNID_PRO|SOL_COM|SOLSAI_PRO|ITSOLSAI_|LOTE$|LOTE_|ALMOX|KIT$|KIT_|ITSOL_COM|SALDO_|ENDERECAMENTO|SEPARACAO|FRACIONAMENTO|ARSENAL|ETIQUETA|INVENTARIO|CONSUMO$)', re.IGNORECASE), '10_FARMACIA_ESTOQUE'),

    # Compras/contratos
    (re.compile(r'^(GCTO_|COMPRA|PEDIDO_COMPRA|ITPED_COMPRA|FORNEC|LICITACAO|COTACAO|COTADOR|CONTRATO|PREGAO|ADESAO_PROCESSO)', re.IGNORECASE), '11_COMPRAS_CONTRATOS'),

    # Faturamento
    (re.compile(r'^(REGFAT|ITREG|REG_REP|FAT_|CONTA$|CONTA_|ITCONTA_|FATURA|VALOR_|VAL_|PROCED_|PRO_FAT|TABPROC|LANCAMENTO_|GLOSA|REPASSE|IT_REPASSE|AGIR_)', re.IGNORECASE), '12_FATURAMENTO_CONTAS'),

    # Financeiro
    (re.compile(r'^(FINANC|TITULO_|MOVFIN|CAIXA|BANCO$|BANCO_|TESOUR|RECEB_|PAGTO_|FLUXO_|CHEQUE|BOLETO|BORDERO|DUPLICATA|PLANO_CONTAS|MODELO_CONTABIL|CCOR_|CUSTO_|CUSTO$|BASE_ORCA|PECA_ORCAMENTARIA|TET_ORC|ORCA_|ORCAMENTO|DRE$|MOEDA$|APOLICE|SEGUROS|ALCADA|ALCA_|RUBRICA)', re.IGNORECASE), '13_FINANCEIRO'),

    # Patrimonio
    (re.compile(r'^(PATRIM|BEM_|BENS$|MANUT_|EQUIP_|CALIBR|ENG_CLIN|DEPRECIA|OFICINA)', re.IGNORECASE), '16_PATRIMONIO_ENGENHARIA'),

    # Agenda/ambulatorio
    (re.compile(r'^(AGENDA$|AGENDA_|AGE_RX|IT_AGENDA|HORARIO_AGENDA|CONSULTA|AMBULAT)', re.IGNORECASE), '03_AMBULATORIO_AGENDAMENTO'),

    # Atendimento
    (re.compile(r'^(ATENDIME|ATESTADO|LEITO$|LEITO_|UNID_INT|INTERNACAO|ALTA_|TRANSF_|MOVIM_LEITO|CENSO|ADMISSAO)', re.IGNORECASE), '02_ATENDIMENTO'),

    # Paciente/cadastro geral
    (re.compile(r'^(PACIENTE|REGISTRO_|CONVENIO|PESSOA|ENDERECO|CONTATO|RESP_|CIDADE$|ESTADO$|PAIS$|UF$|MUNICIPIO|LOCALIDADE|REGIAO$|CID$|CIDO$|CBOR$|TUSS$|CNAE|CNAER|CNO$|CSOSN|CST$|CFOP|NBS$|PROFISSAO|ESPECIALID|ESPECIALIZACOES|CONSELHO|ETNIA|RELIGIAO|NACIONALIDADE|CIDADANIAS)', re.IGNORECASE), '01_CADASTRO_PACIENTE'),

    # Integracao
    (re.compile(r'^(INTEGRA_|IMV_|FUSION_|HL7_|FHIR_|WS_|WEBSERV|API_|DE_PARA|SINCRONIZACAO|EXPORTADOR)', re.IGNORECASE), '18_INTEGRACAO'),

    # Auditoria/seguranca
    (re.compile(r'^(LOG_|AUDIT_|BACKUP_|HIST_|USUARIO|USU_|PERFIL_|PERMISSAO|ACESSO_|NOTIFICACAO)', re.IGNORECASE), '20_SEGURANCA_AUDITORIA'),

    # Configuracao/sistema
    (re.compile(r'^(CONFIG|TIPO_|TIP_|PARAM_|MULTI_EMPRESA|EMPRESA_|UNIDADE$|UNIDADE_|SETOR$|SETOR_|GRUPO$|GRUPO_|GRU_|AGRUP|MOTIVO|MOT_|REGRA$|REGRA_|CLASSIFICACAO|CLASSE$|MODELO_|PRIORIDADE|MENSAG|VERSAO|HOSPITAL$|ESCALA|TURMA$|TURNO$|FERIADO|HABILIDADE|SITE_|BATCH_|DATABASECHANGE|SEMAFORO|GCM_|HGG_|SCR_|CONFIGEST|ACRO_|MES_ANO|TABELAS$|COLUNAS$|COMANDO$|IMP_|IMPRESSAO)', re.IGNORECASE), '21_CONFIGURACAO_SISTEMA'),

    # BI / DW
    (re.compile(r'^(FATO_|FA_|DI_|DIM_|OLAP_|CUBO_)', re.IGNORECASE), '17_BI_DATA_WAREHOUSE'),

    # Portal
    (re.compile(r'^(PORTLET_|PLAN_|SD_|DASHBOARD_|PAINEL_)', re.IGNORECASE), '19_PORTAL_DASHBOARD'),

    # Nutricao/copa (vai para config como suporte hospitalar)
    (re.compile(r'^(COPA$|PRATO|ALIMENTO|NUTRIENTE|DIETA)', re.IGNORECASE), '21_CONFIGURACAO_SISTEMA'),

    # Prestador/equipe
    (re.compile(r'^(PRESTADOR|EQUIPE$|EQUIPE_|FUNCIONAR|FUNCIONARIO|COOPERATIVA|CREDENCIADO|AUTORIZADOR)', re.IGNORECASE), '01_CADASTRO_PACIENTE'),

    # Documentos gerais
    (re.compile(r'^(DOCUMENTO$|DOCUMENTO_|GABARITO|QUESTIONARIO|PERGUNTA|PERGUNTAS)', re.IGNORECASE), '21_CONFIGURACAO_SISTEMA'),

    # Lavanderia/esterilizacao (suporte hospitalar)
    (re.compile(r'^(LAVANDERIA|ESTERILIZACAO|MATERIAL$|MATERIAIS)', re.IGNORECASE), '21_CONFIGURACAO_SISTEMA'),

    # Servico/setor
    (re.compile(r'^(SERVICO$|SER_DIS|ENT_SERV)', re.IGNORECASE), '21_CONFIGURACAO_SISTEMA'),

    # Infeccao hospitalar
    (re.compile(r'^(INFECCAO|INVASIVOS)', re.IGNORECASE), '08_EXAMES_LABORATORIO'),

    # Vacina/imunizacao
    (re.compile(r'^(VACINA)', re.IGNORECASE), '06_PRESCRICAO_MEDICACAO'),

    # Veiculos/transporte
    (re.compile(r'^(MOTORISTA|TRAJETO|VEICULO)', re.IGNORECASE), '21_CONFIGURACAO_SISTEMA'),

    # OSV (Ordem de Servico)
    (re.compile(r'^(OSV_)', re.IGNORECASE), '16_PATRIMONIO_ENGENHARIA'),

    # Mais padroes DBAMV especificos
    # Financeiro extra
    (re.compile(r'^(FINAN|LCTO_|PLANO$|PLANO_|SOLIC_|MOV_CONT|ANO_FECH|FAIXA_|PAGU_|ACRESC_|ALIQUOTA|IMPOSTO|TRIBUT|ICMS|ISS_|PIS_|COFINS|IRRF|IOF_|IPTU)', re.IGNORECASE), '13_FINANCEIRO'),

    # Estoque/movimentacao extra
    (re.compile(r'^(MVTO_|MOV_EST|ROMANEIO|ITEM_ENT|ITEM_SAI|ITEM_TRANS)', re.IGNORECASE), '10_FARMACIA_ESTOQUE'),

    # Prontuario extra
    (re.compile(r'^(PEPWEB|TERMO_|ESTADIAMENTO|METASTASE|TOPOGRAFIA|TRATAMENTO|NOTIF_)', re.IGNORECASE), '05_PRONTUARIO_ELETRONICO'),

    # Lab extra
    (re.compile(r'^(COLETA_|FORMULA$|FORMULA_|ITFORMULA|MANIPULA|ITMANIPULA|COMPON)', re.IGNORECASE), '08_EXAMES_LABORATORIO'),

    # Atendimento extra
    (re.compile(r'^(TROCPAC|ALTAS_|ACIDENTE_)', re.IGNORECASE), '02_ATENDIMENTO'),

    # Integracao extra
    (re.compile(r'^(IMVSAP|DADOS_)', re.IGNORECASE), '18_INTEGRACAO'),

    # Config/sistema extra
    (re.compile(r'^(STATUS_|CONTROLE_|HISTORICO_|CATEGORIA_|REPORTS|TABELA_|ARQUIVO_|PROGRAMA_|ISO_|CRITICA_|IMG_|PROJETO$|PROJETO_|RMI_|TEF_|VISAO_|APLICACAO|AREA$|AREA_|SUB_|TP_|C_|PT_|REG_|CON_|TAB_|ITEM_|IT_)', re.IGNORECASE), '21_CONFIGURACAO_SISTEMA'),

    # Temporarios/backup -> outros
    (re.compile(r'^(TMPMV_|TEMP_|TMP_|BKP_|#TABLEAU|CMP3|CMP4|ANDRE$|TESTE$|LOCKS$)', re.IGNORECASE), '99_OUTROS'),
]

# Fase 3: Palavras-chave no COMENTARIO_TABELA
KEYWORD_RULES = [
    (['paciente', 'cadastro de paciente', 'registro do paciente'], '01_CADASTRO_PACIENTE'),
    (['atendimento', 'internacao', 'internação', 'leito', 'alta do paciente'], '02_ATENDIMENTO'),
    (['agenda', 'agendamento', 'ambulatorio', 'ambulatório', 'consulta'], '03_AMBULATORIO_AGENDAMENTO'),
    (['urgencia', 'urgência', 'emergencia', 'emergência', 'triagem', 'classificacao de risco', 'classificação de risco'], '04_URGENCIA_CLASSIFICACAO_RISCO'),
    (['prontuario', 'prontuário', 'evolucao', 'evolução', 'documento clinico', 'documento clínico'], '05_PRONTUARIO_ELETRONICO'),
    (['prescricao', 'prescrição', 'medicacao', 'medicação', 'medicamento', 'aprazamento'], '06_PRESCRICAO_MEDICACAO'),
    (['cirurgia', 'cirurgico', 'cirúrgico', 'anestesia', 'opme', 'centro cirurgico'], '07_CENTRO_CIRURGICO'),
    (['exame', 'laboratorio', 'laboratório', 'resultado', 'laudo', 'sadt', 'amostra'], '08_EXAMES_LABORATORIO'),
    (['imagem', 'dicom', 'pacs', 'radiolog'], '09_IMAGEM_PACS'),
    (['farmacia', 'farmácia', 'estoque', 'almoxarifado', 'dispensacao', 'dispensação'], '10_FARMACIA_ESTOQUE'),
    (['compra', 'fornecedor', 'contrato', 'licitacao', 'licitação', 'cotacao', 'cotação'], '11_COMPRAS_CONTRATOS'),
    (['faturamento', 'conta medica', 'conta médica', 'procedimento', 'glosa', 'cobranca', 'cobrança'], '12_FATURAMENTO_CONTAS'),
    (['financeiro', 'titulo', 'título', 'pagamento', 'recebimento', 'caixa', 'tesouraria'], '13_FINANCEIRO'),
    (['tiss', 'sus', 'ans', 'regulat', 'nota fiscal'], '14_REGULATORIO_TISS_SUS'),
    (['plano de saude', 'plano de saúde', 'operadora', 'beneficiario', 'beneficiário'], '15_PLANOS_SAUDE'),
    (['patrimonio', 'patrimônio', 'equipamento', 'manutencao', 'manutenção', 'bem patrimonial'], '16_PATRIMONIO_ENGENHARIA'),
    (['fato', 'dimensao', 'dimensão', 'olap', 'cubo', 'data warehouse'], '17_BI_DATA_WAREHOUSE'),
    (['integracao', 'integração', 'webservice', 'api', 'hl7', 'interoperabilidade'], '18_INTEGRACAO'),
    (['portal', 'painel', 'dashboard', 'portlet'], '19_PORTAL_DASHBOARD'),
    (['log', 'auditoria', 'usuario', 'usuário', 'permissao', 'permissão', 'seguranca', 'segurança'], '20_SEGURANCA_AUDITORIA'),
    (['configuracao', 'configuração', 'parametro', 'parâmetro', 'tipo de', 'sistema'], '21_CONFIGURACAO_SISTEMA'),
]


# ============================================================
# ENGINE DE CLASSIFICACAO
# ============================================================
def classify_table(owner, table_name, comment):
    """Classifica uma tabela. Retorna (categoria_id, regra_usada)."""
    owner_upper = str(owner).strip().upper() if pd.notna(owner) else ''
    table_upper = str(table_name).strip().upper() if pd.notna(table_name) else ''
    comment_str = str(comment).strip().lower() if pd.notna(comment) else ''

    # Fase 1: Owner
    if owner_upper in OWNER_RULES:
        return OWNER_RULES[owner_upper], 'OWNER'
    if owner_upper in OWNER_OTHERS:
        return '99_OUTROS', 'OWNER'

    # Fase 2: Prefixo
    for pattern, cat_id in PREFIX_RULES:
        if pattern.match(table_upper):
            return cat_id, 'PREFIXO'

    # Fase 3: Keyword no comentario
    if comment_str and comment_str != 'nan':
        for keywords, cat_id in KEYWORD_RULES:
            if any(kw in comment_str for kw in keywords):
                return cat_id, 'KEYWORD'

    # Fase 4: Default
    return '99_OUTROS', 'DEFAULT'


def classify_all_tables(df):
    """Classifica todas as tabelas unicas e retorna DataFrame com resultados."""
    tables = df.groupby(['OWNER', 'TABLE_NAME']).agg(
        COMENTARIO_TABELA=('COMENTARIO_TABELA', 'first'),
        Qtd_Colunas=('COLUMN_NAME', 'count'),
        ACESSOMV=('ACESSOMV', 'first'),
    ).reset_index()

    classifications = tables.apply(
        lambda row: classify_table(row['OWNER'], row['TABLE_NAME'], row['COMENTARIO_TABELA']),
        axis=1,
    )
    tables['Categoria'] = classifications.apply(lambda x: x[0])
    tables['Regra_Classificacao'] = classifications.apply(lambda x: x[1])
    tables['Nome_Categoria'] = tables['Categoria'].map(
        lambda c: CATEGORIES.get(c, {}).get('name', 'Outros')
    )
    tables['Arquivo'] = tables['Categoria'] + '.xlsx'

    return tables


# ============================================================
# FORMATACAO EXCEL
# ============================================================
def style_worksheet(ws, header_color='1F4E79'):
    """Aplica formatacao ao worksheet: header colorido, filtro, freeze."""
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            val = str(cell.value) if cell.value else ''
            if len(val) > max_len:
                max_len = len(val)
        adjusted = min(max(max_len + 2, 12), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    ws.freeze_panes = 'A2'
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


# ============================================================
# GERACAO DOS ARQUIVOS
# ============================================================
def generate_index_file(classified, output_dir):
    """Gera o arquivo indice 00_INDICE_MV_TABELAS.xlsx."""
    filepath = output_dir / '00_INDICE_MV_TABELAS.xlsx'

    # Aba 1: Categorias
    cat_summary = classified.groupby('Categoria').agg(
        Nome_Categoria=('Nome_Categoria', 'first'),
        Qtd_Tabelas=('TABLE_NAME', 'nunique'),
        Qtd_Colunas=('Qtd_Colunas', 'sum'),
        Owners=('OWNER', lambda x: ', '.join(sorted(x.unique()))),
    ).reset_index()
    cat_summary['Descricao'] = cat_summary['Categoria'].map(
        lambda c: CATEGORIES.get(c, {}).get('description', '')
    )
    cat_summary['Arquivo'] = cat_summary['Categoria'] + '.xlsx'
    cat_summary = cat_summary.rename(columns={'Categoria': 'Codigo_Categoria'})
    cat_summary = cat_summary[['Codigo_Categoria', 'Nome_Categoria', 'Descricao',
                                'Arquivo', 'Qtd_Tabelas', 'Qtd_Colunas', 'Owners']]
    cat_summary = cat_summary.sort_values('Codigo_Categoria')

    # Aba 2: Diretorio de tabelas
    table_dir = classified[['OWNER', 'TABLE_NAME', 'COMENTARIO_TABELA',
                             'Categoria', 'Nome_Categoria', 'Arquivo',
                             'Qtd_Colunas', 'Regra_Classificacao']].copy()
    table_dir = table_dir.sort_values(['Categoria', 'OWNER', 'TABLE_NAME'])

    # Aba 3: Owners
    owner_summary = classified.groupby('OWNER').agg(
        Qtd_Tabelas=('TABLE_NAME', 'nunique'),
        Categorias=('Categoria', lambda x: ', '.join(sorted(x.unique()))),
    ).reset_index().sort_values('Qtd_Tabelas', ascending=False)

    # Aba 4: Estatisticas
    stats = pd.DataFrame({
        'Metrica': ['Total_Tabelas', 'Total_Colunas', 'Total_Owners',
                     'Total_Categorias', 'Tabelas_em_Outros', 'Data_Geracao'],
        'Valor': [
            str(classified['TABLE_NAME'].nunique()),
            str(int(classified['Qtd_Colunas'].sum())),
            str(classified['OWNER'].nunique()),
            str(classified['Categoria'].nunique()),
            str(len(classified[classified['Categoria'] == '99_OUTROS'])),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ],
    })

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        cat_summary.to_excel(writer, sheet_name='Categorias', index=False)
        table_dir.to_excel(writer, sheet_name='Tabelas', index=False)
        owner_summary.to_excel(writer, sheet_name='Owners', index=False)
        stats.to_excel(writer, sheet_name='Estatisticas', index=False)

        style_worksheet(writer.sheets['Categorias'], '1F4E79')
        style_worksheet(writer.sheets['Tabelas'], '2E75B6')
        style_worksheet(writer.sheets['Owners'], '1F4E79')
        style_worksheet(writer.sheets['Estatisticas'], '385723')

    return filepath


def generate_category_files(classified, df_full, output_dir):
    """Gera um arquivo Excel para cada categoria."""
    # Merge da categoria no df completo
    df_with_cat = df_full.merge(
        classified[['OWNER', 'TABLE_NAME', 'Categoria']],
        on=['OWNER', 'TABLE_NAME'],
        how='left',
    )

    files_created = []
    for cat_id, group in df_with_cat.groupby('Categoria'):
        filepath = output_dir / f'{cat_id}.xlsx'

        # Aba 1: Resumo de tabelas
        cat_tables = classified[classified['Categoria'] == cat_id]
        sheet1 = cat_tables[['OWNER', 'TABLE_NAME', 'COMENTARIO_TABELA',
                              'Qtd_Colunas', 'ACESSOMV']].copy()
        sheet1 = sheet1.sort_values(['OWNER', 'TABLE_NAME'])

        # Aba 2: Detalhe das colunas
        col_order = ['OWNER', 'TABLE_NAME', 'COLUMN_NAME', 'DATA_TYPE',
                      'NULLABLE', 'COLUMN_ID', 'COMENTARIO_TABELA',
                      'COMENTARIO_COLUNA', 'ACESSOMV']
        existing_cols = [c for c in col_order if c in group.columns]
        sheet2 = group[existing_cols].copy()
        if 'COLUMN_ID' in sheet2.columns:
            sheet2 = sheet2.sort_values(['OWNER', 'TABLE_NAME', 'COLUMN_ID'])
        else:
            sheet2 = sheet2.sort_values(['OWNER', 'TABLE_NAME'])

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            sheet1.to_excel(writer, sheet_name='Tabelas', index=False)
            sheet2.to_excel(writer, sheet_name='Colunas', index=False)

            style_worksheet(writer.sheets['Tabelas'], '1F4E79')
            style_worksheet(writer.sheets['Colunas'], '2E75B6')

        cat_name = CATEGORIES.get(cat_id, {}).get('name', cat_id)
        print(f'       {filepath.name} - {len(cat_tables)} tabelas - {cat_name}')
        files_created.append(filepath)

    return files_created


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='Separa MV Tabelas.xlsx por area funcional')
    parser.add_argument('--input', default='MV Tabelas.xlsx', help='Arquivo Excel de entrada')
    parser.add_argument('--output', default='output_mv_categorizado', help='Pasta de saida')
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f'[1/5] Carregando {input_path}...')
    df = pd.read_excel(input_path, engine='openpyxl', header=1)

    # Normaliza nomes das colunas
    col_mapping = {}
    for col in df.columns:
        col_clean = str(col).strip().upper()
        if 'OWNER' in col_clean:
            col_mapping[col] = 'OWNER'
        elif 'TABLE_NAME' in col_clean:
            col_mapping[col] = 'TABLE_NAME'
        elif 'COLUMN_NAME' in col_clean:
            col_mapping[col] = 'COLUMN_NAME'
        elif 'DATA_TYPE' in col_clean:
            col_mapping[col] = 'DATA_TYPE'
        elif 'NULLABLE' in col_clean:
            col_mapping[col] = 'NULLABLE'
        elif 'COLUMN_ID' in col_clean:
            col_mapping[col] = 'COLUMN_ID'
        elif 'COMENTARIO_TABELA' in col_clean:
            col_mapping[col] = 'COMENTARIO_TABELA'
        elif 'COMENTARIO_COLUNA' in col_clean:
            col_mapping[col] = 'COMENTARIO_COLUNA'
        elif 'ACESSOMV' in col_clean:
            col_mapping[col] = 'ACESSOMV'
    df.rename(columns=col_mapping, inplace=True)

    # Se header=1 nao funcionou, tenta header=0
    required = ['OWNER', 'TABLE_NAME', 'COLUMN_NAME']
    if not all(c in df.columns for c in required):
        print('       Header na linha 2 nao encontrado, tentando linha 1...')
        df = pd.read_excel(input_path, engine='openpyxl', header=0)
        col_mapping = {}
        for col in df.columns:
            col_clean = str(col).strip().upper()
            if 'OWNER' in col_clean:
                col_mapping[col] = 'OWNER'
            elif 'TABLE_NAME' in col_clean:
                col_mapping[col] = 'TABLE_NAME'
            elif 'COLUMN_NAME' in col_clean:
                col_mapping[col] = 'COLUMN_NAME'
            elif 'DATA_TYPE' in col_clean:
                col_mapping[col] = 'DATA_TYPE'
            elif 'NULLABLE' in col_clean:
                col_mapping[col] = 'NULLABLE'
            elif 'COLUMN_ID' in col_clean:
                col_mapping[col] = 'COLUMN_ID'
            elif 'COMENTARIO_TABELA' in col_clean:
                col_mapping[col] = 'COMENTARIO_TABELA'
            elif 'COMENTARIO_COLUNA' in col_clean:
                col_mapping[col] = 'COMENTARIO_COLUNA'
            elif 'ACESSOMV' in col_clean:
                col_mapping[col] = 'ACESSOMV'
        df.rename(columns=col_mapping, inplace=True)

    # Garante que colunas opcionais existam
    for col in ['COMENTARIO_TABELA', 'COMENTARIO_COLUNA', 'ACESSOMV']:
        if col not in df.columns:
            df[col] = None

    print(f'       Carregadas {len(df):,} linhas, {df["TABLE_NAME"].nunique():,} tabelas unicas')
    print(f'       Colunas detectadas: {list(df.columns)}')

    print('[2/5] Classificando tabelas...')
    classified = classify_all_tables(df)

    # Resumo por categoria
    for cat_id in sorted(classified['Categoria'].unique()):
        count = len(classified[classified['Categoria'] == cat_id])
        cat_name = CATEGORIES.get(cat_id, {}).get('name', cat_id)
        print(f'       {cat_id}: {count:,} tabelas - {cat_name}')

    print('[3/5] Gerando arquivo indice...')
    idx_file = generate_index_file(classified, output_dir)
    print(f'       Criado: {idx_file.name}')

    print('[4/5] Gerando arquivos por categoria...')
    files = generate_category_files(classified, df, output_dir)

    total_files = len(files) + 1  # +1 para o indice
    print(f'[5/5] Concluido! {total_files} arquivos gerados em {output_dir}/')

    # Avisos
    outros = classified[classified['Categoria'] == '99_OUTROS']
    if len(outros) > 0:
        pct = len(outros) / len(classified) * 100
        print(f'\n  AVISO: {len(outros)} tabelas ({pct:.1f}%) ficaram em 99_OUTROS.')
        print('  Revise esse arquivo e ajuste as regras se necessario.')

    # Distribuicao das regras
    print('\n  Distribuicao por tipo de regra:')
    for regra, count in classified['Regra_Classificacao'].value_counts().items():
        print(f'       {regra}: {count:,} tabelas')


if __name__ == '__main__':
    main()
